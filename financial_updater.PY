# -*- coding: utf-8 -*-
import yfinance as yf
import pandas as pd
import time
import datetime
import os
from pathlib import Path
import random
import numpy as np
import logging
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from email.header import Header
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from cryptography.fernet import Fernet
import sqlite3
from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.chart import LineChart, Reference, Series

# --- CONFIGURACIÓN GENERAL ---
TICKERS = ['AAPL', 'GOOGL', 'MSFT', 'NVDA', 'AMZN', 'META', 'TSLA',  # Tech Grandes
           '^GSPC', 'BTC-USD',  # Indices/Crypto
           'BRK-B', 'BAC', 'JNJ',  # Financieras/Salud/Conglomerado
           'INTC', 'CSCO', 'AMD',  # Tech Otras
           'DIS', 'XOM', 'WMT', 'PG', 'KO']  # Consumo/Energía/Otros

# --- DEFINICIÓN DE GRUPOS PARA GRÁFICOS (¡AJUSTAR SEGÚN PREFERENCIA!) ---
TICKER_GROUPS = {
    "Tech Gigantes": ['AAPL', 'GOOGL', 'MSFT', 'NVDA', 'AMZN', 'META', 'TSLA'],
    "Índice y Crypto": ['^GSPC', 'BTC-USD'],
    "Semiconductores y Redes": ['INTC', 'CSCO', 'AMD'],
    "Finanzas, Salud, Conglom.": ['BRK-B', 'BAC', 'JNJ'],
    "Consumo, Energía, Entreten.": ['DIS', 'XOM', 'WMT', 'PG', 'KO']
}
all_grouped_tickers = set(t for group in TICKER_GROUPS.values() for t in group)
if set(TICKERS) != all_grouped_tickers: logging.warning(f"Discrepancia TICKERS/TICKER_GROUPS! Faltan: {set(TICKERS) - all_grouped_tickers}, Sobran: {all_grouped_tickers - set(TICKERS)}")

OUTPUT_DIR = Path("./Financial_Data")
LOGS_DIR = Path("./Logs")
OUTPUT_FILE_BASE_NAME = "Dynamic Financial Data.xlsx"
OUTPUT_FILE = OUTPUT_DIR / OUTPUT_FILE_BASE_NAME
LOG_FILE = LOGS_DIR / "financial_updater.log"

# --- INTERVALOS DE ACTUALIZACIÓN RECOMENDADOS ---
FREQ_UPDATE_INTERVAL_MINUTES = 15
FULL_UPDATE_INTERVAL_HOURS = 6

# --- PARÁMETROS DE ANÁLISIS ---
HISTORY_PERIOD = '1y'
SMA_SHORT = 50
SMA_LONG = 200

# --- CONFIGURACIÓN SERVICIOS EXTERNOS (¡RELLENAR!) ---
EMAIL_SENDER = "tu_email@gmail.com"
EMAIL_PASSWORD = "tu_contraseña_o_contraseña_app"
EMAIL_RECIPIENT = "destinatario@example.com"
GOOGLE_SHEETS_CREDENTIALS_FILE = "credentials.json"
GOOGLE_SHEETS_BOOK_NAME = "Nombre del Libro"
GOOGLE_SHEETS_WORKSHEET_NAME = "Resumen General"

# --- CREACIÓN DE DIRECTORIOS ---
def create_directories():
    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        LOGS_DIR.mkdir(parents=True, exist_ok=True)
        print(f"Directorios: {OUTPUT_DIR}, {LOGS_DIR}")
    except OSError as e:
        print(f"Error directorios: {e}")
        raise

# --- CONFIGURACIÓN DE LOGS (Consola + Archivo) ---
create_directories()
log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
root_logger = logging.getLogger()
root_logger.setLevel(logging.INFO)
for handler in root_logger.handlers[:]:
    root_logger.removeHandler(handler)
try:
    file_handler = logging.FileHandler(LOG_FILE, encoding='utf-8')
    file_handler.setFormatter(log_formatter)
    root_logger.addHandler(file_handler)
except Exception as e:
    print(f"Error log archivo: {e}")
try:
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(log_formatter)
    root_logger.addHandler(console_handler)
except Exception as e:
    print(f"Error log consola: {e}")

# --- CLASE PRINCIPAL ---
class FinancialDataUpdater:
    def __init__(self, tickers, output_dir, freq_interval_min, full_interval_hr=None):
        self.tickers = tickers
        self.output_dir = output_dir
        self.freq_update_interval = freq_interval_min
        self.full_update_interval = full_interval_hr
        self.output_file_path = self.output_dir / OUTPUT_FILE_BASE_NAME

    # --- FUNCIONES DE OBTENCIÓN DE DATOS ---
    def fetch_full_financial_data(self):
        logging.info(f"[FULL] Iniciando obtención COMPLETA datos...")
        summary_data_list, all_history_data, financial_sheets, news_data = [], {}, {}, []
        history_df_close, sma_short_df, sma_long_df, rsi_df = pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        try:
            logging.info(f"[FULL] Descargando historial ({HISTORY_PERIOD})...")
            history_df_full = yf.download(self.tickers, period=HISTORY_PERIOD, progress=False, timeout=60, auto_adjust=False, group_by='ticker')
            logging.info("[FULL] Descarga historial OK.")
            if not history_df_full.empty:
                if isinstance(history_df_full.columns, pd.MultiIndex):
                    level1 = history_df_full.columns.get_level_values(1)
                    col = 'Adj Close' if 'Adj Close' in level1 else ('Close' if 'Close' in level1 else None)
                    history_df_close = history_df_full.loc[:, pd.IndexSlice[:, col]] if col else pd.DataFrame()
                    history_df_close.columns = history_df_close.columns.droplevel(1) if col else []
                else:
                    history_df_close = history_df_full[['Adj Close']] if 'Adj Close' in history_df_full.columns else history_df_full[['Close']] if 'Close' in history_df_full.columns else pd.DataFrame()
                if not history_df_close.empty:
                    all_history_data['Adj Close'] = history_df_close
                    logging.info(f"[FULL] Historial procesado. Shape: {history_df_close.shape}")
                    numeric_cols = history_df_close.select_dtypes(include=np.number).columns
                    if len(numeric_cols) > 0:
                        logging.info("[FULL] Calculando SMAs/RSI...")
                        sma_short_df = history_df_close[numeric_cols].rolling(window=SMA_SHORT, min_periods=SMA_SHORT).mean()
                        sma_long_df = history_df_close[numeric_cols].rolling(window=SMA_LONG, min_periods=SMA_LONG).mean()
                        rsi_df = self.calculate_rsi(history_df_close)
                        logging.info("[FULL] SMAs/RSI OK.")
            else:
                logging.warning("[FULL] Historial descargado vacío.")
        except Exception as e:
            logging.error(f"[FULL] Error historial: {e}", exc_info=True)
        logging.info("[FULL] Obteniendo datos individuales...")
        yf_tickers = yf.Tickers(self.tickers)
        for ticker_symbol in self.tickers:
            logging.info(f"[FULL] Procesando: {ticker_symbol}")
            try:
                ticker_obj = yf_tickers.tickers.get(ticker_symbol)
                ticker_info = ticker_obj.info if ticker_obj else None
                if not ticker_info or ticker_info.get('quoteType') == 'EMPTY':
                    logging.warning(f"[FULL] Info inválida {ticker_symbol}.")
                    continue
                def get_value(d, k, f=1, t=(int, float)):
                    v = d.get(k)
                    return v * f if v is not None and isinstance(v, t) else np.nan
                summary_dict = {
                    'Ticker': ticker_symbol,
                    'Nombre': ticker_info.get('shortName', 'N/A'),
                    'Precio Actual': get_value(ticker_info, 'currentPrice') or get_value(ticker_info, 'regularMarketPrice') or get_value(ticker_info, 'previousClose'),
                    'Cambio Hoy (%)': get_value(ticker_info, 'regularMarketChangePercent', 100),
                    'Máx Hoy': get_value(ticker_info, 'dayHigh'),
                    'Mín Hoy': get_value(ticker_info, 'dayLow'),
                    'Volumen': get_value(ticker_info, 'regularMarketVolume', t=(int, float, str)),
                    'Capitalización Mercado': get_value(ticker_info, 'marketCap', t=(int, float, str)),
                    'PER': get_value(ticker_info, 'trailingPE'),
                    'EPS Trail': get_value(ticker_info, 'trailingEps'),
                    'EPS Fwd': get_value(ticker_info, 'forwardEps'),
                    'P/B Ratio': get_value(ticker_info, 'priceToBook'),
                    'Beta': get_value(ticker_info, 'beta'),
                    'Rend. Dividendo (%)': get_value(ticker_info, 'dividendYield', 100),
                    'Tasa Dividendo': get_value(ticker_info, 'dividendRate'),
                    'Fecha Ex-Dividendo': datetime.datetime.fromtimestamp(ticker_info.get('exDividendDate')).strftime('%Y-%m-%d') if ticker_info.get('exDividendDate') else 'N/A',
                    'Máx 52 Sem': get_value(ticker_info, 'fiftyTwoWeekHigh'),
                    'Mín 52 Sem': get_value(ticker_info, 'fiftyTwoWeekLow'),
                    'Target Precio Medio': get_value(ticker_info, 'targetMeanPrice'),
                    'Target Precio Alto': get_value(ticker_info, 'targetHighPrice'),
                    'Target Precio Bajo': get_value(ticker_info, 'targetLowPrice'),
                    'Recom. Media': ticker_info.get('recommendationMean', 'N/A'),
                    'Recom. Clave': ticker_info.get('recommendationKey', 'N/A'),
                    'Última Actualización Info': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                sma_s = sma_short_df[ticker_symbol].iloc[-1] if not sma_short_df.empty and ticker_symbol in sma_short_df and not sma_short_df[ticker_symbol].isnull().all() else np.nan
                sma_l = sma_long_df[ticker_symbol].iloc[-1] if not sma_long_df.empty and ticker_symbol in sma_long_df and not sma_long_df[ticker_symbol].isnull().all() else np.nan
                rsi = rsi_df[ticker_symbol].iloc[-1] if not rsi_df.empty and ticker_symbol in rsi_df and not rsi_df[ticker_symbol].isnull().all() else np.nan
                summary_dict.update({f'SMA {SMA_SHORT}': sma_s, f'SMA {SMA_LONG}': sma_l, 'RSI': rsi})
                h52, l52, curr = summary_dict['Máx 52 Sem'], summary_dict['Mín 52 Sem'], summary_dict['Precio Actual']
                summary_dict['% Rango 52 Sem'] = ((curr - l52) / (h52 - l52)) * 100 if pd.notna(curr) and pd.notna(h52) and pd.notna(l52) and h52 > l52 else np.nan
                try:
                    fin, bs, cf = ticker_obj.financials, ticker_obj.balance_sheet, ticker_obj.cashflow
                    financial_sheets.update({f"{ticker_symbol}_Financials": fin} if fin is not None and not fin.empty else {})
                    financial_sheets.update({f"{ticker_symbol}_BalanceSheet": bs} if bs is not None and not bs.empty else {})
                    financial_sheets.update({f"{ticker_symbol}_Cashflow": cf} if cf is not None and not cf.empty else {})
                except Exception as e:
                    logging.warning(f"[FULL] Financieros {ticker_symbol}: {e}")
                try:
                    news = ticker_obj.news
                    news_data.extend([{'Ticker': ticker_symbol, 'Título': item.get('title'), 'Publicador': item.get('publisher'), 'Enlace': item.get('link'), 'Tipo': item.get('type'), 'Fecha': datetime.datetime.fromtimestamp(item.get('providerPublishTime')).strftime('%Y-%m-%d %H:%M:%S') if item.get('providerPublishTime') else 'N/A'} for item in news] if news else [])
                except Exception as e:
                    logging.warning(f"[FULL] Noticias {ticker_symbol}: {e}")
                summary_data_list.append(summary_dict)
                time.sleep(random.uniform(0.2, 0.6))
            except Exception as e:
                logging.error(f"[FULL] Error procesando {ticker_symbol}: {e}", exc_info=True)
        df_summary = pd.DataFrame(summary_data_list)
        df_news = pd.DataFrame(news_data)
        if not df_summary.empty:
            cols = ['Ticker', 'Nombre', 'Precio Actual', 'Cambio Hoy (%)', f'SMA {SMA_SHORT}', f'SMA {SMA_LONG}', 'RSI', 'Máx Hoy', 'Mín Hoy', 'Volumen', 'Capitalización Mercado', 'PER', 'EPS Trail', 'EPS Fwd', 'P/B Ratio', 'Beta', 'Rend. Dividendo (%)', 'Tasa Dividendo', 'Fecha Ex-Dividendo', 'Máx 52 Sem', 'Mín 52 Sem', '% Rango 52 Sem', 'Target Precio Medio', 'Target Precio Alto', 'Target Precio Bajo', 'Recom. Media', 'Recom. Clave', 'Última Actualización Info']
            df_summary = df_summary.reindex(columns=cols).fillna('N/A')
        if not df_news.empty:
            df_news = df_news.sort_values(by='Fecha', ascending=False)
        data_dict = {'summary': df_summary, 'history': all_history_data, 'financials': financial_sheets, 'news': df_news}
        logging.info("[FULL] Obtención COMPLETA datos finalizada.")
        return data_dict

    def fetch_live_data(self):
        logging.info(f"[LIVE] Iniciando obtención RÁPIDA datos...")
        live_data_list = []
        yf_tickers = yf.Tickers(self.tickers)
        for ticker_symbol in self.tickers:
            try:
                ticker_obj = yf_tickers.tickers.get(ticker_symbol)
                ticker_info = ticker_obj.info if ticker_obj else None
                if not ticker_info or ticker_info.get('quoteType') == 'EMPTY':
                    continue
                price = ticker_info.get('currentPrice') or ticker_info.get('regularMarketPrice') or ticker_info.get('previousClose')
                change_pct = ticker_info.get('regularMarketChangePercent')
                volume = ticker_info.get('regularMarketVolume')
                live_dict = {
                    'Ticker': ticker_symbol,
                    'Precio Live': price if price is not None else np.nan,
                    'Cambio % Live': (change_pct * 100) if isinstance(change_pct, (int, float)) else np.nan,
                    'Volumen Live': volume if isinstance(volume, (int, float)) else np.nan,
                    'Timestamp Live': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                live_data_list.append(live_dict)
                time.sleep(random.uniform(0.1, 0.3))
            except Exception as e:
                logging.error(f"[LIVE] Error {ticker_symbol}: {e}", exc_info=False)
        if not live_data_list:
            logging.warning("[LIVE] No datos en vivo.")
        logging.info("[LIVE] Obtención RÁPIDA datos finalizada.")
        return pd.DataFrame(live_data_list)

    # --- FUNCIONES DE CÁLCULO ---
    def calculate_rsi(self, prices, window=14):
        if prices.empty or len(prices) < window + 1:
            logging.warning(f"Datos insuficientes RSI {window}.")
            return pd.DataFrame(index=prices.index, columns=prices.columns)
        delta = prices.diff()
        gain = delta.where(delta > 0, 0.0)
        loss = -delta.where(delta < 0, 0.0)
        avg_gain = gain.ewm(com=window - 1, min_periods=window).mean()
        avg_loss = loss.ewm(com=window - 1, min_periods=window).mean()
        rs = avg_gain / avg_loss
        rsi = 100.0 - (100.0 / (1.0 + rs))
        return rsi

    # --- FUNCIONES DE ESCRITURA Y FORMATO ---
    def write_full_data_to_excel(self, data_dict):
        logging.info(f"[FULL] Iniciando escritura COMPLETA: {self.output_file_path}")
        filepath = self.output_file_path
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='w', datetime_format='YYYY-MM-DD HH:MM:SS') as writer:
                if 'summary' in data_dict and not data_dict['summary'].empty:
                    data_dict['summary'].to_excel(writer, sheet_name='Resumen General', index=False)
                if 'history' in data_dict and 'Adj Close' in data_dict['history'] and not data_dict['history']['Adj Close'].empty:
                    hist_df = data_dict['history']['Adj Close'].reset_index().rename(columns={data_dict['history']['Adj Close'].index.name or 'index': 'Fecha'})
                    hist_df.to_excel(writer, sheet_name='Historial_Adj_Close', index=False)
                if 'financials' in data_dict and data_dict['financials']:
                    for sheet_name, df in data_dict['financials'].items():
                        if df is not None and not df.empty:
                            df.reset_index().rename(columns={df.index.name or 'index': 'Metrica'}).to_excel(writer, sheet_name=sheet_name, index=False)
                if 'news' in data_dict and not data_dict['news'].empty:
                    data_dict['news'].to_excel(writer, sheet_name='Noticias Recientes', index=False)
            logging.info("[FULL] Escritura inicial COMPLETA OK.")
            return self.apply_excel_formatting(filepath, data_dict)
        except Exception as e:
            logging.error(f"[FULL] Error escribiendo Excel COMPLETO {filepath}: {e}", exc_info=True)
            return False

    def write_live_data_to_excel(self, live_df):
        if live_df.empty:
            logging.warning("[LIVE] DF vivo vacío.")
            return False
        sheet_name = "Live Data"
        logging.info(f"[LIVE] Escribiendo hoja '{sheet_name}' en: {self.output_file_path}")
        try:
            if not self.output_file_path.exists():
                workbook = Workbook()
                logging.warning(f"[LIVE] Archivo {self.output_file_path} no existe. Creando.")
                del workbook['Sheet']
            else:
                workbook = load_workbook(self.output_file_path)
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
            worksheet = workbook.create_sheet(sheet_name)
            headers = live_df.columns.tolist()
            worksheet.append(headers)
            for r_idx, row in enumerate(live_df.itertuples(index=False), 2):
                for c_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    cell.value = None if pd.isna(value) else value
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
                        fmt = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        if '%' in headers[c_idx-1]:
                            fmt = numbers.FORMAT_PERCENTAGE_00
                        cell.number_format = fmt
                    elif 'Timestamp' in headers[c_idx-1]:
                        cell.alignment = Alignment(horizontal='center')
            for i, col_letter in enumerate(get_column_letter(j+1) for j in range(live_df.shape[1])):
                worksheet.column_dimensions[col_letter].width = 18
            workbook.save(self.output_file_path)
            logging.info(f"[LIVE] Hoja '{sheet_name}' actualizada.")
            return True
        except Exception as e:
            logging.error(f"[LIVE] Error escribiendo hoja '{sheet_name}': {e}", exc_info=True)
            return False

    def apply_excel_formatting(self, filename, data_dict):
        logging.info(f"[FULL] Aplicando formato y gráficos a: {filename}")
        try:
            workbook = load_workbook(filename)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            fmt_price = numbers.FORMAT_CURRENCY_USD_SIMPLE
            fmt_price_no_symbol = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            fmt_percent = numbers.FORMAT_PERCENTAGE_00
            fmt_volume = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            fmt_ratio = '0.00'
            fmt_date = 'yyyy-mm-dd'
            fmt_integer = '0'
            link_font = Font(underline="single", color="0563C1")
            wrap_alignment = Alignment(vertical='top', wrap_text=True)

            def apply_standard_formatting(sheet, column_formats, column_widths, freeze_panes_coord='A2'):
                if not sheet:
                    return
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                for r_idx in range(2, sheet.max_row + 1):
                    for c_letter, (fmt, align) in column_formats.items():
                        cell = sheet[f"{c_letter}{r_idx}"]
                        cell.alignment = align if align else cell.alignment
                        if fmt and isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                            cell.number_format = fmt
                            if fmt == fmt_percent and abs(cell.value) > 1.5:
                                cell.value = cell.value / 100.0
                for c_letter, width in column_widths.items():
                    sheet.column_dimensions[c_letter].width = width
                if freeze_panes_coord:
                    sheet.freeze_panes = sheet[freeze_panes_coord]

            if 'Resumen General' in workbook.sheetnames:
                sheet = workbook['Resumen General']
                fmts = {
                    'A': (None, left_align),
                    'B': (None, left_align),
                    'C': (fmt_price, right_align),
                    'D': (fmt_percent, right_align),
                    'E': (fmt_price_no_symbol, right_align),
                    'F': (fmt_price_no_symbol, right_align),
                    'G': (fmt_integer, right_align),
                    'H': (fmt_price, right_align),
                    'I': (fmt_price, right_align),
                    'J': (fmt_volume, right_align),
                    'K': (fmt_volume, right_align),
                    'L': (fmt_ratio, right_align),
                    'M': (fmt_ratio, right_align),
                    'N': (fmt_ratio, right_align),
                    'O': (fmt_ratio, right_align),
                    'P': (fmt_ratio, right_align),
                    'Q': (fmt_percent, right_align),
                    'R': (fmt_ratio, right_align),
                    'S': (fmt_date, center_align),
                    'T': (fmt_price, right_align),
                    'U': (fmt_price, right_align),
                    'V': (fmt_percent, right_align),
                    'W': (fmt_price, right_align),
                    'X': (fmt_price, right_align),
                    'Y': (fmt_price, right_align),
                    'Z': (fmt_ratio, center_align),
                    'AA': (None, center_align),
                    'AB': (None, center_align)
                }
                widths = {
                    'A': 10, 'B': 25, 'C': 14, 'D': 12, 'E': 14, 'F': 14, 'G': 8, 'H': 14, 'I': 14, 'J': 20, 'K': 25, 'L': 10, 'M': 10, 'N': 10, 'O': 10, 'P': 8, 'Q': 12, 'R': 10, 'S': 14, 'T': 14, 'U': 14, 'V': 12, 'W': 14, 'X': 14, 'Y': 14, 'Z': 12, 'AA': 15, 'AB': 20
                }
                apply_standard_formatting(sheet, fmts, widths, 'B2')
                logging.info("[FULL] Formato 'Resumen General' OK.")

            if 'financials' in data_dict:
                for sheet_name in data_dict['financials'].keys():
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        fmts = {}
                        widths = {'A': 35}
                        for i in range(2, sheet.max_column + 1):
                            fmts[get_column_letter(i)] = (fmt_volume, right_align)
                            widths[get_column_letter(i)] = 18
                        apply_standard_formatting(sheet, fmts, widths, 'B2')
                        logging.info(f"[FULL] Formato '{sheet_name}' OK.")

            if 'Noticias Recientes' in workbook.sheetnames:
                sheet = workbook['Noticias Recientes']
                fmts = {
                    'A': (None, center_align),
                    'B': (None, left_align),
                    'C': (None, left_align),
                    'D': (None, left_align),
                    'E': (None, center_align),
                    'F': (None, center_align)
                }
                widths = {'A': 10, 'B': 70, 'C': 25, 'D': 40, 'E': 15, 'F': 20}
                apply_standard_formatting(sheet, fmts, widths, 'A2')
                for r in range(2, sheet.max_row + 1):
                    for c in range(1, sheet.max_column + 1):
                        sheet.cell(r, c).alignment = wrap_alignment
                    link_c = sheet.cell(r, 4)
                    title_c = sheet.cell(r, 2)
                    if link_c.value and isinstance(link_c.value, str) and link_c.value.startswith('http'):
                        link_c.hyperlink = link_c.value
                        link_c.value = title_c.value[:50] + "..." if len(title_c.value) > 50 else "Link"
                        link_c.font = link_font
                logging.info("[FULL] Formato 'Noticias Recientes' OK.")

            if 'Historial_Adj_Close' in workbook.sheetnames:
                sheet = workbook['Historial_Adj_Close']
                fmts = {'A': (fmt_date, center_align)}
                widths = {'A': 20}
                for i in range(2, sheet.max_column + 1):
                    fmts[get_column_letter(i)] = (fmt_price_no_symbol, right_align)
                    widths[get_column_letter(i)] = 15
                apply_standard_formatting(sheet, fmts, widths, 'B2')
                logging.info("[FULL] Formato 'Historial_Adj_Close' OK.")

            df_historial = None
            df_normalizado = None
            if 'history' in data_dict and 'Adj Close' in data_dict['history'] and not data_dict['history']['Adj Close'].empty:
                df_historial = data_dict['history']['Adj Close'].copy()
                df_historial.index = pd.to_datetime(df_historial.index, errors='coerce')
                if not isinstance(df_historial.index, pd.DatetimeIndex):
                    df_historial = None
            if df_historial is not None:
                try:
                    first_valid_idx = df_historial.apply(pd.Series.first_valid_index)
                    df_first_values = pd.Series([df_historial.at[idx, col] for col, idx in first_valid_idx.items()], index=df_historial.columns)
                    df_normalizado = df_historial.apply(lambda col: (col / df_first_values[col.name]) * 100 if col.name in df_first_values and pd.notna(df_first_values[col.name]) and df_first_values[col.name] != 0 else col, axis=0)
                    df_normalizado.fillna(method='ffill', inplace=True)
                    logging.info("[FULL] Datos normalizados OK.")
                except Exception as e:
                    logging.error(f"[FULL] Error normalizando: {e}", exc_info=True)
            if df_normalizado is not None and not df_normalizado.empty:
                temp_sheet_name = "Temp_Norm_Data_For_Chart"
                temp_sheet = workbook.create_sheet(temp_sheet_name) if temp_sheet_name not in workbook.sheetnames else workbook[temp_sheet_name]
                temp_sheet.delete_rows(1, temp_sheet.max_row)
                temp_sheet.append(['Fecha'] + df_normalizado.columns.tolist())
                for ts, row in df_normalizado.iterrows():
                    temp_sheet.append([ts] + row.tolist())
                temp_sheet.column_dimensions['A'].width = 20
                logging.info(f"[FULL] Hoja temporal '{temp_sheet_name}' OK.")
                chart_sheet_name = "Gráficos"
                chart_sheet = workbook.create_sheet(chart_sheet_name) if chart_sheet_name not in workbook.sheetnames else workbook[chart_sheet_name]
                chart_sheet._charts = []
                current_chart_row = 1
                chart_height_approx_rows = 18
                temp_headers = [cell.value for cell in temp_sheet[1]]
                logging.info("[FULL] Creando gráficos por grupo...")
                for group_name, group_tickers in TICKER_GROUPS.items():
                    logging.info(f"[FULL] -> Grupo: {group_name}")
                    chart = LineChart()
                    chart.title = f"Rendimiento Normalizado - {group_name}"
                    chart.style = 12
                    chart.x_axis.title = "Fecha"
                    chart.y_axis.title = "Rendimiento (%)"
                    chart.y_axis.number_format = '0"%"'
                    chart.height = 10
                    chart.width = 20
                    chart.legend.position = 'b'
                    dates_ref = Reference(temp_sheet, min_col=1, min_row=2, max_row=temp_sheet.max_row)
                    chart.set_categories(dates_ref)
                    ticker_found = False
                    for ticker in group_tickers:
                        try:
                            if ticker in temp_headers:
                                col_idx = temp_headers.index(ticker) + 1
                                data_col_ref = Reference(temp_sheet, min_col=col_idx, min_row=2, max_row=temp_sheet.max_row)
                                series = Series(values=data_col_ref, title=ticker)
                                chart.series.append(series)
                                ticker_found = True
                            else:
                                logging.warning(f"[FULL] Ticker '{ticker}' grupo '{group_name}' no encontrado.")
                        except Exception as e:
                            logging.error(f"[FULL] Error añadiendo serie '{ticker}' gráfico '{group_name}': {e}")
                    if ticker_found:
                        anchor = f"A{current_chart_row}"
                        chart_sheet.add_chart(chart, anchor)
                        logging.info(f"[FULL] Gráfico '{group_name}' añadido en {anchor}.")
                        current_chart_row += chart_height_approx_rows
                    else:
                        logging.warning(f"[FULL] Omitiendo gráfico '{group_name}', sin tickers.")
                # if temp_sheet_name in workbook.sheetnames:
                #     del workbook[temp_sheet_name]
                #     logging.info(f"[FULL] Hoja temporal '{temp_sheet_name}' eliminada.")
            else:
                logging.warning("[FULL] No se crearán gráficos.")
            workbook.save(filename)
            logging.info(f"[FULL] Archivo Excel '{filename}' guardado OK.")
            return True
        except Exception as e:
            logging.error(f"[FULL] Error crítico formateo/gráfico: {e}", exc_info=True)
            return False

    # --- FUNCIONES SERVICIOS EXTERNOS Y DB ---
    def send_email(self, subject, body, to_addr, attachment_path=None):
        logging.info(f"Intentando enviar correo a: {to_addr}")
        if not EMAIL_SENDER or not EMAIL_PASSWORD or not to_addr:
            logging.warning("Faltan credenciales email.")
            return False
        try:
            msg = MIMEMultipart()
            msg['Subject'] = Header(subject, 'utf-8').encode()
            msg['From'] = EMAIL_SENDER
            msg['To'] = to_addr
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            if attachment_path and Path(attachment_path).exists():
                filename = Path(attachment_path).name
                with open(attachment_path, "rb") as fil:
                    part = MIMEApplication(fil.read(), Name=filename)
                part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                msg.attach(part)
                logging.info(f"Adjunto: {filename}")
            elif attachment_path:
                logging.warning(f"Adjunto no existe: {attachment_path}")
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login(EMAIL_SENDER, EMAIL_PASSWORD)
                server.send_message(msg)
            logging.info("Correo enviado.")
        except smtplib.SMTPAuthenticationError as e:
            logging.error(f"Error Auth SMTP: {e}.", exc_info=True)
            return False
        except Exception as e:
            logging.error(f"Error enviando correo: {e}", exc_info=True)
            return False
        else:
            return True

    def update_google_sheets(self, data_dict):
        logging.info("Intentando actualizar GSheets...")
        if not GOOGLE_SHEETS_CREDENTIALS_FILE or not GOOGLE_SHEETS_BOOK_NAME or not GOOGLE_SHEETS_WORKSHEET_NAME:
            logging.warning("Falta config GSheets.")
            return False
        if not Path(GOOGLE_SHEETS_CREDENTIALS_FILE).exists():
            logging.error(f"No credenciales: {GOOGLE_SHEETS_CREDENTIALS_FILE}")
            return False
        if 'summary' not in data_dict or data_dict['summary'].empty:
            logging.warning("No datos resumen GSheets.")
            return False
        try:
            scope = [
                "https://spreadsheets.google.com/feeds",
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive.file",
                "https://www.googleapis.com/auth/drive"
            ]
            creds = ServiceAccountCredentials.from_json_keyfile_name(GOOGLE_SHEETS_CREDENTIALS_FILE, scope)
            client = gspread.authorize(creds)
            logging.info(f"Abriendo GSheet: '{GOOGLE_SHEETS_BOOK_NAME}' -> Hoja: '{GOOGLE_SHEETS_WORKSHEET_NAME}'")
            sheet = client.open(GOOGLE_SHEETS_BOOK_NAME).worksheet(GOOGLE_SHEETS_WORKSHEET_NAME)
            df_upload = data_dict['summary'].fillna('').copy()
            logging.info("Limpiando y subiendo a GSheets...")
            sheet.clear()
            sheet.update([df_upload.columns.values.tolist()] + df_upload.values.tolist(), value_input_option='USER_ENTERED')
            logging.info("GSheets actualizado.")
        except gspread.exceptions.SpreadsheetNotFound:
            logging.error(f"GSheet no encontrado: '{GOOGLE_SHEETS_BOOK_NAME}'")
            return False
        except gspread.exceptions.WorksheetNotFound:
            logging.error(f"Hoja no encontrada: '{GOOGLE_SHEETS_WORKSHEET_NAME}'")
            return False
        except Exception as e:
            logging.error(f"Error GSheets: {e}", exc_info=True)
            return False
        else:
            return True

    def store_in_db(self, dataframe, table_name, db_name="financial_data.db"):
        if dataframe.empty:
            logging.warning(f"DF vacío tabla '{table_name}'.")
            return False
        logging.info(f"Almacenando en SQLite: '{db_name}', Tabla: '{table_name}'")
        try:
            db_path = Path(db_name)
            db_path.parent.mkdir(parents=True, exist_ok=True)
            conn = sqlite3.connect(db_name)
            dataframe.to_sql(table_name, conn, if_exists="replace", index=False)
            conn.close()
            logging.info(f"Datos SQLite tabla '{table_name}' OK.")
        except Exception as e:
            logging.error(f"Error SQLite tabla '{table_name}': {e}", exc_info=True)
            return False
        else:
            return True

    # --- MÉTODOS DE TRABAJO (JOBS) ---
    def job_full_update(self):
        logging.info("="*70 + f"\n[FULL] INICIANDO JOB COMPLETO - {datetime.datetime.now()}")
        start_time = time.time()
        success = False
        try:
            data_dict = self.fetch_full_financial_data()
            if data_dict and not data_dict['summary'].empty:
                success = self.write_full_data_to_excel(data_dict)
                if success:
                    self.update_google_sheets(data_dict)
                    self.store_in_db(data_dict['summary'], 'summary')
        except Exception as e:
            logging.critical(f"[FULL] Error CRÍTICO job completo: {e}", exc_info=True)
        finally:
            duration = time.time() - start_time
            logging.info(f"[FULL] FINALIZANDO JOB COMPLETO - Duración: {duration:.2f} seg - Éxito: {success}\n" + "="*70 + "\n")

    def job_frequent_update(self):
        logging.info(f"[LIVE] Iniciando Job Rápido - {datetime.datetime.now()}")
        start_time = time.time()
        success = False
        try:
            live_df = self.fetch_live_data()
            success = self.write_live_data_to_excel(live_df) if not live_df.empty else False
        except Exception as e:
            logging.error(f"[LIVE] Error job rápido: {e}", exc_info=True)
        finally:
            duration = time.time() - start_time
            logging.info(f"[LIVE] Finalizando Job Rápido - Duración: {duration:.2f} seg - Éxito Escritura: {success}\n")

    # --- INICIO DEL SCHEDULER ---
    def start_scheduler(self):
        logging.info("Configurando scheduler...")
        scheduler = BackgroundScheduler(daemon=True)
        if self.full_update_interval:
            logging.info(f"Programando Job COMPLETO cada {self.full_update_interval} horas.")
            scheduler.add_job(self.job_full_update, 'interval', hours=self.full_update_interval, misfire_grace_time=300)
        else:
            logging.warning("Intervalo Job COMPLETO no configurado.")
        logging.info(f"Programando Job RÁPIDO cada {self.freq_update_interval} minutos.")
        scheduler.add_job(self.job_frequent_update, 'interval', minutes=self.freq_update_interval, misfire_grace_time=60)
        scheduler.start()
        logging.info(f"Scheduler iniciado. Ctrl+C para detener.")
        try:
            while True:
                time.sleep(5)
        except (KeyboardInterrupt, SystemExit):
            logging.info("Interrupción recibida. Deteniendo scheduler...")
            scheduler.shutdown()
            logging.info("Scheduler detenido limpiamente. Saliendo del script.")

# --- PUNTO DE ENTRADA ---
if __name__ == "__main__":
    logging.info("Iniciando script...")
    try:
        updater = FinancialDataUpdater(
            tickers=TICKERS,
            output_dir=OUTPUT_DIR,
            freq_interval_min=FREQ_UPDATE_INTERVAL_MINUTES,
            full_interval_hr=FULL_UPDATE_INTERVAL_HOURS
        )
        logging.info("Ejecutando primer job COMPLETO...")
        updater.job_full_update()
        updater.start_scheduler()
    except Exception as e:
        logging.critical(f"Error fatal inicialización: {e}", exc_info=True)
        print(f"Error fatal, revisa log: {LOG_FILE}.")
    finally:
        logging.info("Script principal terminado.")